"""
БГ Финансови Отчети — Streamlit App
Генерира Баланс и ОПР в НСИ формат от оборотна ведомост.
"""

import streamlit as st
import pandas as pd
import numpy as np
import json
import os
import io
from pathlib import Path

# ──────────────────────────────────────────────────────────────────────────────
# NSI BALANCE SHEET TEMPLATE
# ──────────────────────────────────────────────────────────────────────────────
# Each row: (code, label, indent_level, is_total, sign)
# sign = +1 → add positive amounts; sign used for net calculation
# is_total = True → row is computed as sum of preceding rows in its section

BALANCE_TEMPLATE = [
    # ── АКТИВ ──────────────────────────────────────────────────────────────
    ("AKTIV",        "АКТИВ",                                                    0, True,  1),
    ("1000",         "I. Нетекущи (дълготрайни) активи",                        1, True,  1),
    ("1100",         "1. Нематериални активи",                                   2, True,  1),
    ("1110",         "Концесии, патенти, лицензи, търг. марки и подобни права",  3, False, 1),
    ("1120",         "Права върху сгради и терени",                              3, False, 1),
    ("1130",         "Продукти от развойна дейност",                             3, False, 1),
    ("1140",         "Репутация",                                                3, False, 1),
    ("1150",         "Аванси и нематериални активи в процес на придобиване",     3, False, 1),
    ("1190",         "Други нематериални активи",                                3, False, 1),
    ("1200",         "2. Имоти, машини, съоръжения и оборудване",               2, True,  1),
    ("1210",         "Земи и терени",                                            3, False, 1),
    ("1220",         "Сгради и конструкции",                                     3, False, 1),
    ("1230",         "Машини, производствено оборудване и апаратура",           3, False, 1),
    ("1240",         "Транспортни средства",                                     3, False, 1),
    ("1250",         "Офис оборудване",                                          3, False, 1),
    ("1260",         "Аванси и МСО в процес на придобиване",                    3, False, 1),
    ("1290",         "Други имоти, машини, съоръжения и оборудване",            3, False, 1),
    ("1300",         "3. Инвестиционни имоти",                                   2, False, 1),
    ("1400",         "4. Биологични активи",                                     2, False, 1),
    ("1500",         "5. Дългосрочни финансови активи",                         2, True,  1),
    ("1510",         "Акции и дялове в дъщерни предприятия",                    3, False, 1),
    ("1520",         "Акции и дялове в асоциирани предприятия",                 3, False, 1),
    ("1530",         "Дългосрочни инвестиции",                                  3, False, 1),
    ("1540",         "Заеми, предоставени на свързани предприятия",             3, False, 1),
    ("1550",         "Дългосрочни вземания",                                    3, False, 1),
    ("1590",         "Други дългосрочни финансови активи",                      3, False, 1),
    ("1600",         "6. Търговска репутация",                                   2, False, 1),
    ("1700",         "7. Активи по отсрочени данъци",                           2, False, 1),
    ("2000",         "II. Текущи (краткотрайни) активи",                        1, True,  1),
    ("2100",         "1. Материални запаси",                                    2, True,  1),
    ("2110",         "Суровини и материали",                                    3, False, 1),
    ("2120",         "Незавършено производство",                                3, False, 1),
    ("2130",         "Готова продукция",                                        3, False, 1),
    ("2140",         "Стоки",                                                   3, False, 1),
    ("2190",         "Други материални запаси",                                 3, False, 1),
    ("2200",         "2. Вземания",                                             2, True,  1),
    ("2210",         "Търговски вземания",                                      3, False, 1),
    ("2220",         "Вземания от свързани предприятия",                        3, False, 1),
    ("2230",         "Вземания по данъци и осигуровки",                        3, False, 1),
    ("2240",         "Вземания от персонала",                                   3, False, 1),
    ("2290",         "Други вземания",                                          3, False, 1),
    ("2300",         "3. Краткосрочни инвестиции",                              2, True,  1),
    ("2310",         "Акции и дялове",                                          3, False, 1),
    ("2390",         "Други краткосрочни инвестиции",                           3, False, 1),
    ("2400",         "4. Парични средства",                                     2, True,  1),
    ("2410",         "Каса",                                                    3, False, 1),
    ("2420",         "Разплащателни сметки в лева",                             3, False, 1),
    ("2430",         "Разплащателни сметки във валута",                         3, False, 1),
    ("2490",         "Други парични средства",                                  3, False, 1),
    ("2500",         "5. Разходи за бъдещи периоди",                           2, False, 1),
    ("SUMA_A",       "СУМА НА АКТИВА",                                          0, True,  1),
    # ── ПАСИВ ──────────────────────────────────────────────────────────────
    ("PASIV",        "ПАСИВ",                                                   0, True,  1),
    ("3000",         "I. Собствен капитал",                                     1, True,  1),
    ("3100",         "1. Основен капитал",                                      2, False, 1),
    ("3200",         "2. Резерви",                                              2, True,  1),
    ("3210",         "Законови резерви",                                        3, False, 1),
    ("3220",         "Резерви от последваща оценка",                            3, False, 1),
    ("3290",         "Други резерви",                                           3, False, 1),
    ("3300",         "3. Финансов резултат",                                    2, True,  1),
    ("3310",         "Неразпределена печалба / Непокрита загуба от минали год.", 3, False, 1),
    ("3320",         "Печалба / Загуба за текущата година",                    3, False, 1),
    ("4000",         "II. Провизии и сходни задължения",                        1, False, 1),
    ("5000",         "III. Дългосрочни задължения",                            1, True,  1),
    ("5100",         "Задължения към кредитни институции",                      2, False, 1),
    ("5200",         "Задължения по финансов лизинг",                          2, False, 1),
    ("5300",         "Задължения към свързани предприятия",                     2, False, 1),
    ("5900",         "Други дългосрочни задължения",                            2, False, 1),
    ("6000",         "IV. Текущи задължения",                                  1, True,  1),
    ("6100",         "Задължения към кредитни институции",                      2, False, 1),
    ("6200",         "Задължения по финансов лизинг",                          2, False, 1),
    ("6300",         "Търговски задължения",                                    2, False, 1),
    ("6400",         "Задължения към свързани предприятия",                     2, False, 1),
    ("6500",         "Задължения към персонала",                                2, False, 1),
    ("6600",         "Задължения по данъци и осигуровки",                      2, False, 1),
    ("6900",         "Други текущи задължения",                                 2, False, 1),
    ("7000",         "V. Приходи за бъдещи периоди",                           1, False, 1),
    ("SUMA_P",       "СУМА НА ПАСИВА",                                         0, True,  1),
]

# Total row definitions: code → list of detail codes that sum into it
BALANCE_TOTALS = {
    "1100": ["1110","1120","1130","1140","1150","1190"],
    "1200": ["1210","1220","1230","1240","1250","1260","1290"],
    "1500": ["1510","1520","1530","1540","1550","1590"],
    "1000": ["1100","1200","1300","1400","1500","1600","1700"],
    "2100": ["2110","2120","2130","2140","2190"],
    "2200": ["2210","2220","2230","2240","2290"],
    "2300": ["2310","2390"],
    "2400": ["2410","2420","2430","2490"],
    "2000": ["2100","2200","2300","2400","2500"],
    "SUMA_A": ["1000","2000"],
    "3200": ["3210","3220","3290"],
    "3300": ["3310","3320"],
    "3000": ["3100","3200","3300"],
    "5000": ["5100","5200","5300","5900"],
    "6000": ["6100","6200","6300","6400","6500","6600","6900"],
    "SUMA_P": ["3000","4000","5000","6000","7000"],
    "AKTIV": ["SUMA_A"],
    "PASIV": ["SUMA_P"],
}

# ──────────────────────────────────────────────────────────────────────────────
# NSI OPR TEMPLATE
# ──────────────────────────────────────────────────────────────────────────────
OPR_TEMPLATE = [
    ("10100",  "I. Нетни приходи от продажби",                                0, False, 1),
    ("10110",  "в т.ч. от продажби на продукция",                             1, False, 1),
    ("10120",  "в т.ч. от продажби на стоки",                                 1, False, 1),
    ("10130",  "в т.ч. от продажби на услуги",                                1, False, 1),
    ("10200",  "II. Изменение на запасите от продукция и незавършено пр-во",  0, False, 1),
    ("10300",  "III. Разходи за придобиване на активи за собствена употреба", 0, False, 1),
    ("10400",  "IV. Други приходи от дейността",                              0, False, 1),
    ("10500",  "V. Разходи за суровини, материали и външни услуги",           0, False, -1),
    ("10510",  "в т.ч. разходи за суровини и материали",                     1, False, -1),
    ("10520",  "в т.ч. разходи за външни услуги",                             1, False, -1),
    ("10600",  "VI. Разходи за персонала",                                    0, False, -1),
    ("10610",  "в т.ч. разходи за заплати и възнаграждения",                 1, False, -1),
    ("10620",  "в т.ч. разходи за осигуровки",                               1, False, -1),
    ("10700",  "VII. Разходи за амортизация и обезценка",                    0, False, -1),
    ("10800",  "VIII. Други разходи от дейността",                           0, False, -1),
    ("EBIT",   "Печалба/Загуба от оперативна дейност",                        0, True,  1),
    ("10900",  "IX. Приходи от участия и други финансови инструменти",        0, False, 1),
    ("11000",  "X. Приходи от лихви",                                         0, False, 1),
    ("11100",  "XI. Разходи за лихви",                                        0, False, -1),
    ("11200",  "XII. Други финансови приходи",                                0, False, 1),
    ("11300",  "XIII. Други финансови разходи",                               0, False, -1),
    ("EBT",    "Печалба/Загуба преди данъци",                                 0, True,  1),
    ("11400",  "XIV. Разходи за данъци",                                      0, False, -1),
    ("ФР",     "Печалба/Загуба за периода",                                   0, True,  1),
    ("11500",  "в т.ч. за продължаващи дейности",                            1, False, 1),
    ("11600",  "в т.ч. за преустановени дейности",                           1, False, 1),
]

OPR_TOTALS = {
    "EBIT": ["10100","10200","10300","10400","10500","10600","10700","10800"],
    "EBT":  ["EBIT","10900","11000","11100","11200","11300"],
    "ФР":   ["EBT","11400"],
}

# ──────────────────────────────────────────────────────────────────────────────
# PREFIX → NSI MAPPING TABLE
# Each prefix (string): (balance_nsi_code, opr_nsi_code, acct_type, label)
# acct_type: "A"=asset, "L"=liability/equity, "E"=expense, "R"=revenue, "X"=ignore
# balance_nsi_code: which balance row to add this account to (None if not balance)
# opr_nsi_code:     which OPR row to add this account to (None if not OPR)
# ──────────────────────────────────────────────────────────────────────────────
PREFIX_MAP = {
    # ── 1xx  Нематериални активи ──────────────────────────────────────────
    "101": ("1110", None, "A", "Нематериални активи — концесии, патенти"),
    "102": ("1110", None, "A", "Нематериални активи — разработки"),
    "104": ("1130", None, "A", "Продукти от развойна дейност"),
    "105": ("1140", None, "A", "Репутация"),
    "107": ("1150", None, "A", "Аванси за нематериални активи"),
    "108": ("1190", None, "A", "Други нематериални активи"),
    # Амортизации на НМА — контра-актив, намалява родителя
    "241": ("1110", None, "A", "Амортизация — нематериални активи"),

    # ── 2xx  Дълготрайни материални активи ───────────────────────────────
    "201": ("1210", None, "A", "Земи и терени"),
    "202": ("1220", None, "A", "Сгради"),
    "203": ("1230", None, "A", "Машини и оборудване"),
    "204": ("1230", None, "A", "Производствено оборудване"),
    "205": ("1240", None, "A", "Транспортни средства"),
    "206": ("1250", None, "A", "Офис оборудване"),
    "207": ("1260", None, "A", "Разходи за придобиване на ДА"),
    "209": ("1290", None, "A", "Други ДМА"),
    "21":  ("1260", None, "A", "ДМА в процес на придобиване"),
    # Амортизации на ДМА
    "242": ("1220", None, "A", "Амортизация — сгради"),
    "243": ("1230", None, "A", "Амортизация — машини и оборудване"),
    "244": ("1230", None, "A", "Амортизация — производствено оборудване"),
    "245": ("1240", None, "A", "Амортизация — транспортни средства"),
    "246": ("1250", None, "A", "Амортизация — офис оборудване"),
    "249": ("1290", None, "A", "Амортизация — други ДМА"),

    # ── 22x / 23x  Дългосрочни финансови активи ──────────────────────────
    "221": ("1510", None, "A", "Акции в дъщерни дружества"),
    "222": ("1520", None, "A", "Акции в асоциирани дружества"),
    "223": ("1530", None, "A", "Дългосрочни инвестиции"),
    "228": ("1540", None, "A", "Заеми на свързани лица — дълготрайни"),
    "229": ("1550", None, "A", "Дългосрочни вземания"),

    # ── 27x  Инвестиционни имоти ──────────────────────────────────────────
    "271": ("1300", None, "A", "Инвестиционни имоти"),
    "247": ("1300", None, "A", "Амортизация — инвестиционни имоти"),

    # ── 28x  Биологични активи ────────────────────────────────────────────
    "281": ("1400", None, "A", "Биологични активи"),

    # ── 3xx  Материални запаси ────────────────────────────────────────────
    "301": ("2110", None, "A", "Суровини и материали"),
    "302": ("2110", None, "A", "Спомагателни материали"),
    "303": ("2130", None, "A", "Готова продукция"),
    "304": ("2120", None, "A", "Незавършено производство"),
    "305": ("2140", None, "A", "Стоки"),
    "306": ("2190", None, "A", "Амбалаж"),
    "307": ("2190", None, "A", "Млади животни"),
    "309": ("2190", None, "A", "Други материали"),
    "321": ("2110", None, "A", "Доставки на материали (транзит)"),
    "322": ("2140", None, "A", "Доставки на стоки (транзит)"),

    # ── 4xx  Вземания / задължения ────────────────────────────────────────
    # Клиенти (вземания)
    "411": ("2210", None, "A", "Клиенти — търговски вземания"),
    "412": ("2210", None, "A", "Клиенти по рекламации"),
    "413": ("2210", None, "A", "Клиенти — съмнителни"),
    "414": ("2220", None, "A", "Вземания от свързани предприятия"),
    "415": ("2220", None, "A", "Вземания от асоциирани предприятия"),
    "416": ("2290", None, "A", "Вземания по застраховки"),
    "419": ("2290", None, "A", "Клиенти — аванси"),
    # Доставчици (задължения)
    "401": ("6300", None, "L", "Доставчици — търговски задължения"),
    "402": ("6300", None, "L", "Доставчици — аванси получени"),
    "403": ("6300", None, "L", "Доставчици — задължения по рекламации"),
    "404": ("5300", None, "L", "Задължения към свързани предприятия"),
    "405": ("5300", None, "L", "Задължения към асоциирани предприятия"),
    # Персонал
    "421": ("6500", None, "L", "Задължения към персонала"),
    "422": ("6500", None, "L", "Задължения за отпуски"),
    "423": ("6500", None, "L", "Дялово участие на работниците"),
    "424": ("2240", None, "A", "Вземания от подотчетни лица"),
    "425": ("6500", None, "L", "Задължения към подотчетни лица"),
    # Осигуровки / данъци
    "450": ("6600", None, "L", "Задължения за ДДС"),
    "451": ("6600", None, "L", "ДДС — данъчен кредит"),
    "452": ("6600", None, "L", "ДДС — начислен данък"),
    "453": ("6600", None, "L", "Разчети с НОИ — осигуровки"),
    "454": ("6600", None, "L", "Разчети за данък върху доходите на персонала"),
    "455": ("6600", None, "L", "Разчети за данъци и такси"),
    "456": ("2230", None, "A", "Вземания по данъци и осигуровки"),
    "457": ("6600", None, "L", "Корпоративен данък"),
    "459": ("6600", None, "L", "Разчети по данъци — други"),
    # Разчети с акционери / собственици
    "461": ("6900", None, "L", "Задължения към акционери/съдружници"),
    "462": ("2290", None, "A", "Вземания от акционери/съдружници"),
    # Краткосрочни заеми
    "151": ("5100", None, "L", "Задължения към банки — дългосрочни"),
    "152": ("5200", None, "L", "Финансов лизинг — дългосрочен"),
    "159": ("5900", None, "L", "Дългосрочни задължения — други"),
    "431": ("6100", None, "L", "Задължения към банки — краткосрочни"),
    "432": ("6200", None, "L", "Финансов лизинг — краткосрочен"),
    "439": ("6900", None, "L", "Краткосрочни заеми — други"),
    # Разчети с бюджет
    "441": ("2230", None, "A", "Вземания от бюджета — ДДС за възстановяване"),
    "442": ("6600", None, "L", "Задължения за ДДС за внасяне"),
    # Разни вземания / задължения
    "493": ("2290", None, "A", "Разни дебитори"),
    "494": ("6900", None, "L", "Разни кредитори"),
    "495": ("6900", None, "L", "Задължения по гаранции"),
    "496": ("6900", None, "L", "Задължения за гаранционни обезпечения"),
    "497": ("6900", None, "L", "Приходи за бъдещи периоди — текущи"),
    "498": ("2500", None, "A", "Разходи за бъдещи периоди"),
    "499": ("7000", None, "L", "Приходи за бъдещи периоди — дълготрайни"),

    # ── 5xx  Парични средства ─────────────────────────────────────────────
    "501": ("2410", None, "A", "Каса в лева"),
    "502": ("2410", None, "A", "Каса във валута"),
    "503": ("2420", None, "A", "Разплащателна сметка в лева"),
    "504": ("2430", None, "A", "Разплащателна сметка във валута"),
    "505": ("2490", None, "A", "Акредитиви"),
    "506": ("2490", None, "A", "Набирателни сметки"),
    "507": ("2300", None, "A", "Краткосрочни инвестиции — депозити"),
    "509": ("2490", None, "A", "Парични средства в път"),

    # ── Собствен капитал ──────────────────────────────────────────────────
    "101a": ("3100", None, "L", "Основен капитал"),
    "102a": ("3200", None, "L", "Резервен капитал"),
    "111": ("3100", None, "L", "Основен капитал"),
    "112": ("3210", None, "L", "Законови резерви"),
    "113": ("3220", None, "L", "Резерви от последваща оценка"),
    "114": ("3290", None, "L", "Допълнителни резерви"),
    "115": ("3290", None, "L", "Целеви резерви"),
    "119": ("3290", None, "L", "Други резерви"),
    "121": ("3310", None, "L", "Неразпределена печалба"),
    "122": ("3310", None, "L", "Непокрита загуба"),
    "123": ("3320", None, "L", "Печалба за текущата година"),
    "124": ("3320", None, "L", "Загуба за текущата година"),
    "125": ("4000", None, "L", "Провизии"),
    "126": ("4000", None, "L", "Приходи за бъдещи периоди — дългосрочни"),

    # ── 6xx  Разходи ─────────────────────────────────────────────────────
    "601": (None, "10510", "E", "Разходи за суровини и материали"),
    "602": (None, "10520", "E", "Разходи за външни услуги"),
    "603": (None, "10600", "E", "Разходи за амортизация"),
    "604": (None, "10600", "E", "Разходи за заплати"),
    "605": (None, "10620", "E", "Разходи за осигуровки"),
    "606": (None, "10800", "E", "Разходи за данъци, такси, акцизи"),
    "607": (None, "11300", "E", "Разходи по финансови операции"),
    "608": (None, "10800", "E", "Разходи от обезценка на активи"),
    "609": (None, "10800", "E", "Други разходи от дейността"),
    "614": (None, "11100", "E", "Разходи за лихви"),
    "615": (None, "11300", "E", "Отрицателни курсови разлики"),
    "616": (None, "11300", "E", "Разходи от операции с финансови инструменти"),
    "619": (None, "11300", "E", "Други финансови разходи"),
    "621": (None, "11400", "E", "Разходи за данъци от печалбата"),
    "651": (None, "10800", "E", "Разходи от отписване на вземания"),
    "659": (None, "10800", "E", "Разходи от извънредни операции"),

    # ── 7xx  Приходи ─────────────────────────────────────────────────────
    "701": (None, "10130", "R", "Приходи от продажби на услуги"),
    "702": (None, "10120", "R", "Приходи от продажби на стоки"),
    "703": (None, "10110", "R", "Приходи от продажби на продукция"),
    "704": (None, "10110", "R", "Приходи от продажби на материали"),
    "705": (None, "10100", "R", "Приходи от продажби на активи"),
    "706": (None, "10400", "R", "Приходи от финансирания"),
    "709": (None, "10400", "R", "Други приходи от дейността"),
    "714": (None, "11000", "R", "Приходи от лихви"),
    "715": (None, "11200", "R", "Положителни курсови разлики"),
    "716": (None, "11200", "R", "Приходи от операции с финансови инструменти"),
    "718": (None, "10900", "R", "Приходи от участия"),
    "719": (None, "11200", "R", "Други финансови приходи"),
    "721": (None, "10300", "R", "Приходи от придобиване на активи за собствена употреба"),
    "751": (None, "10400", "R", "Приходи от отписване на задължения"),
    "759": (None, "10400", "R", "Извънредни приходи"),
}

# ──────────────────────────────────────────────────────────────────────────────
# HELPER: suggest NSI code based on account prefix
# ──────────────────────────────────────────────────────────────────────────────

def suggest_nsi(acct_num: str, acct_name: str = ""):
    """
    Returns (balance_code, opr_code, acct_type, suggested_label) for an account.
    Tries longest-prefix match first.
    """
    s = str(acct_num).strip()
    for length in (4, 3, 2):
        prefix = s[:length]
        if prefix in PREFIX_MAP:
            return PREFIX_MAP[prefix]
    # Fallback by first digit
    first = s[:1]
    fallbacks = {
        "1": ("1590", None, "A", "Дълготрайни активи — неразпознати"),
        "2": ("1290", None, "A", "ДМА — неразпознати"),
        "3": ("2190", None, "A", "Материални запаси — неразпознати"),
        "4": ("2290", None, "A", "Вземания/задължения — неразпознати"),
        "5": ("2490", None, "A", "Парични средства — неразпознати"),
        "6": (None, "10800", "E", "Разходи — неразпознати"),
        "7": (None, "10400", "R", "Приходи — неразпознати"),
    }
    return fallbacks.get(first, (None, None, "X", "Неразпознато"))


def net_balance(row: pd.Series, ks_dt_col: str, ks_kt_col: str) -> float:
    """Net balance: КС дт − КС кт (positive = debit balance)."""
    dt = float(row.get(ks_dt_col, 0) or 0)
    kt = float(row.get(ks_kt_col, 0) or 0)
    return dt - kt


# ──────────────────────────────────────────────────────────────────────────────
# COMPUTE BALANCE SHEET
# ──────────────────────────────────────────────────────────────────────────────

def compute_balance(tb: pd.DataFrame, mapping: pd.DataFrame,
                    ks_dt_col: str, ks_kt_col: str):
    """
    Aggregate mapped accounts into Balance sheet rows.
    Returns a dict {nsi_code: amount}.
    """
    amounts = {}
    for _, row in tb.iterrows():
        acct = str(row.get("_acct_num", "")).strip()
        net = net_balance(row, ks_dt_col, ks_kt_col)
        # Find mapping for this account
        m = mapping[mapping["account"] == acct]
        if m.empty:
            continue
        bal_code = m.iloc[0].get("bal_code", None)
        acct_type = m.iloc[0].get("type", "X")
        if not bal_code or pd.isna(bal_code):
            continue
        bal_code = str(bal_code)
        # Assets: positive net = debit balance → positive
        # Liabilities/Equity: negative net = credit balance → positive
        if acct_type == "A":
            val = net        # debit-normal
        else:
            val = -net       # credit-normal

        amounts[bal_code] = amounts.get(bal_code, 0.0) + val

    # Fill totals bottom-up
    def fill_total(code):
        children = BALANCE_TOTALS.get(code, [])
        if not children:
            return amounts.get(code, 0.0)
        total = sum(fill_total(c) for c in children)
        amounts[code] = total
        return total

    for code in BALANCE_TOTALS:
        fill_total(code)

    return amounts


# ──────────────────────────────────────────────────────────────────────────────
# COMPUTE OPR
# ──────────────────────────────────────────────────────────────────────────────

def compute_opr(tb: pd.DataFrame, mapping: pd.DataFrame,
                ks_dt_col: str, ks_kt_col: str,
                ob_dt_col: str, ob_kt_col: str):
    """
    Aggregate mapped accounts into OPR rows.
    For pre-closing TB: use closing balance of 6xx/7xx.
    For post-closing TB (6xx/7xx closing ≈ 0): use turnover.
    Returns a dict {nsi_code: amount}.
    """
    amounts = {}
    for _, row in tb.iterrows():
        acct = str(row.get("_acct_num", "")).strip()
        m = mapping[mapping["account"] == acct]
        if m.empty:
            continue
        opr_code = m.iloc[0].get("opr_code", None)
        acct_type = m.iloc[0].get("type", "X")
        if not opr_code or pd.isna(opr_code):
            continue
        opr_code = str(opr_code)

        # Determine closing vs turnover
        ks_net = abs(net_balance(row, ks_dt_col, ks_kt_col))
        ob_dt = float(row.get(ob_dt_col, 0) or 0)
        ob_kt = float(row.get(ob_kt_col, 0) or 0)
        ob_net = abs(ob_dt - ob_kt)

        # Use turnover if closing is near zero (post-closing TB)
        if ks_net < 1.0 and ob_net > 1.0:
            raw = ob_dt if acct_type == "E" else ob_kt
        else:
            net = net_balance(row, ks_dt_col, ks_kt_col)
            raw = net if acct_type == "E" else -net

        amounts[opr_code] = amounts.get(opr_code, 0.0) + abs(raw)

    # Fill totals
    def fill_total(code):
        children = OPR_TOTALS.get(code, [])
        if not children:
            return amounts.get(code, 0.0)
        # EBIT sums revenues minus expenses
        opr_row = {r[0]: r for r in OPR_TEMPLATE}
        total = 0.0
        for c in children:
            c_sign = opr_row[c][4] if c in opr_row else 1
            total += c_sign * fill_total(c)
        amounts[code] = total
        return total

    for code in OPR_TOTALS:
        fill_total(code)

    return amounts


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────────────────────

def build_excel(bal_amounts: dict, opr_amounts: dict, company_name: str = "") -> bytes:
    """Build Excel workbook with Balance and OPR sheets."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    # ── Balance sheet ──────────────────────────────────────────────────────
    ws_bal = wb.active
    ws_bal.title = "Баланс"

    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
    TOTAL1_FILL  = PatternFill("solid", fgColor="BDD7EE")
    TOTAL2_FILL  = PatternFill("solid", fgColor="DDEBF7")
    TOTAL0_FILL  = PatternFill("solid", fgColor="2E75B6")
    THIN = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    BOLD_FONT  = Font(name="Calibri", bold=True, size=10)
    NORM_FONT  = Font(name="Calibri", size=10)

    # Title
    ws_bal.merge_cells("A1:C1")
    ws_bal["A1"] = f"БАЛАНС{' — ' + company_name if company_name else ''}"
    ws_bal["A1"].font = Font(name="Calibri", bold=True, size=14)
    ws_bal["A1"].alignment = Alignment(horizontal="center")

    ws_bal.merge_cells("A2:C2")
    ws_bal["A2"] = "в хиляди лева"
    ws_bal["A2"].alignment = Alignment(horizontal="center")
    ws_bal["A2"].font = Font(name="Calibri", italic=True, size=9)

    headers = ["Показател", "Код", "Сума (хил. лв.)"]
    for ci, h in enumerate(headers, 1):
        c = ws_bal.cell(row=3, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN

    row_num = 4
    for code, label, indent, is_total, _sign in BALANCE_TEMPLATE:
        amount = bal_amounts.get(code, 0.0)
        indent_str = "  " * indent

        c_label = ws_bal.cell(row=row_num, column=1, value=indent_str + label)
        c_code  = ws_bal.cell(row=row_num, column=2, value=code if code not in ("AKTIV","PASIV","SUMA_A","SUMA_P") else "")
        c_amt   = ws_bal.cell(row=row_num, column=3, value=round(amount / 1000, 3) if amount else None)

        c_amt.number_format = '#,##0.000'
        c_amt.alignment = Alignment(horizontal="right")

        if code in ("AKTIV", "PASIV"):
            for c in (c_label, c_code, c_amt):
                c.fill = TOTAL0_FILL
                c.font = WHITE_FONT
                c.border = THIN
        elif code in ("SUMA_A", "SUMA_P") or (is_total and indent == 0):
            for c in (c_label, c_code, c_amt):
                c.fill = TOTAL1_FILL
                c.font = BOLD_FONT
                c.border = THIN
        elif is_total:
            for c in (c_label, c_code, c_amt):
                c.fill = TOTAL2_FILL
                c.font = BOLD_FONT
                c.border = THIN
        else:
            for c in (c_label, c_code, c_amt):
                c.font = NORM_FONT
                c.border = THIN

        row_num += 1

    ws_bal.column_dimensions["A"].width = 60
    ws_bal.column_dimensions["B"].width = 10
    ws_bal.column_dimensions["C"].width = 18
    ws_bal.freeze_panes = "A4"

    # ── OPR sheet ──────────────────────────────────────────────────────────
    ws_opr = wb.create_sheet("ОПР")
    ws_opr.merge_cells("A1:C1")
    ws_opr["A1"] = f"ОТЧЕТ ЗА ПРИХОДИТЕ И РАЗХОДИТЕ{' — ' + company_name if company_name else ''}"
    ws_opr["A1"].font = Font(name="Calibri", bold=True, size=14)
    ws_opr["A1"].alignment = Alignment(horizontal="center")
    ws_opr.merge_cells("A2:C2")
    ws_opr["A2"] = "в хиляди лева"
    ws_opr["A2"].alignment = Alignment(horizontal="center")
    ws_opr["A2"].font = Font(name="Calibri", italic=True, size=9)

    for ci, h in enumerate(headers, 1):
        c = ws_opr.cell(row=3, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN

    row_num = 4
    for code, label, indent, is_total, sign in OPR_TEMPLATE:
        amount = opr_amounts.get(code, 0.0)
        indent_str = "  " * indent

        c_label = ws_opr.cell(row=row_num, column=1, value=indent_str + label)
        c_code  = ws_opr.cell(row=row_num, column=2, value=code if not code.startswith("Ф") and code not in ("EBIT","EBT") else "")
        # For total rows show net; for detail rows show absolute
        display_val = amount if amount else None
        c_amt = ws_opr.cell(row=row_num, column=3, value=round(display_val / 1000, 3) if display_val else None)
        c_amt.number_format = '#,##0.000'
        c_amt.alignment = Alignment(horizontal="right")

        if code in ("ФР", "EBT", "EBIT"):
            for c in (c_label, c_code, c_amt):
                c.fill = TOTAL1_FILL
                c.font = BOLD_FONT
                c.border = THIN
        elif is_total:
            for c in (c_label, c_code, c_amt):
                c.fill = TOTAL2_FILL
                c.font = BOLD_FONT
                c.border = THIN
        else:
            for c in (c_label, c_code, c_amt):
                c.font = NORM_FONT
                c.border = THIN
        row_num += 1

    ws_opr.column_dimensions["A"].width = 65
    ws_opr.column_dimensions["B"].width = 10
    ws_opr.column_dimensions["C"].width = 18
    ws_opr.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ──────────────────────────────────────────────────────────────────────────────
# MAPPING SAVE / LOAD
# ──────────────────────────────────────────────────────────────────────────────
MAPPINGS_DIR = Path("mappings")

def list_mapping_files():
    MAPPINGS_DIR.mkdir(exist_ok=True)
    return sorted(MAPPINGS_DIR.glob("*.json"))

def save_mapping(name: str, df: pd.DataFrame):
    MAPPINGS_DIR.mkdir(exist_ok=True)
    path = MAPPINGS_DIR / f"{name}.json"
    df.to_json(path, orient="records", force_ascii=False, indent=2)

def load_mapping(path) -> pd.DataFrame:
    return pd.read_json(path, orient="records")


# ──────────────────────────────────────────────────────────────────────────────
# COLUMN DETECTION
# ──────────────────────────────────────────────────────────────────────────────

KNOWN_COL_PATTERNS = {
    "acct_num":  ["номер", "сметка №", "№", "acct", "account no", "сч.", "сч. №", "код"],
    "acct_name": ["наименование", "сметка", "name", "account name", "описание"],
    "ns_dt":     ["нс дт", "нс_дт", "os_dt", "нс-дт", "начално дт", "нач. дт", "opening dt"],
    "ns_kt":     ["нс кт", "нс_кт", "os_kt", "нс-кт", "начално кт", "нач. кт", "opening kt"],
    "ob_dt":     ["об дт", "об_дт", "turnover_dt", "обороти дт", "оборот дт", "period dt"],
    "ob_kt":     ["об кт", "об_кт", "turnover_kt", "обороти кт", "оборот кт", "period kt"],
    "ks_dt":     ["кс дт", "кс_дт", "closing_dt", "крайно дт", "кр. дт", "end dt", "ending dt"],
    "ks_kt":     ["кс кт", "кс_кт", "closing_kt", "крайно кт", "кр. кт", "end kt", "ending kt"],
}

def auto_detect_columns(cols):
    """Try to auto-detect TB columns from header names."""
    result = {}
    col_lower = {c: c.lower().strip() for c in cols}
    for key, patterns in KNOWN_COL_PATTERNS.items():
        for col, low in col_lower.items():
            for pat in patterns:
                if pat in low:
                    result[key] = col
                    break
            if key in result:
                break
    return result


# ──────────────────────────────────────────────────────────────────────────────
# STREAMLIT APP
# ──────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="БГ Финансови Отчети",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Custom CSS
st.markdown("""
<style>
    .main-title { font-size: 2rem; font-weight: 700; color: #1F4E79; margin-bottom: 0; }
    .sub-title  { color: #666; margin-top: 0; margin-bottom: 1.5rem; }
    .step-header { font-size: 1.15rem; font-weight: 600; color: #2E75B6;
                   border-left: 4px solid #2E75B6; padding-left: 10px; margin-top: 1rem; }
    .success-box { background: #E8F5E9; border-left: 4px solid #4CAF50;
                   padding: 10px 15px; border-radius: 4px; }
    .warn-box    { background: #FFF3E0; border-left: 4px solid #FF9800;
                   padding: 10px 15px; border-radius: 4px; }
</style>
""", unsafe_allow_html=True)

st.markdown('<p class="main-title">📊 БГ Финансови Отчети</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">Генератор на Баланс и ОПР в НСИ формат от оборотна ведомост</p>',
            unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Настройки")
    company_name = st.text_input("Наименование на фирмата", value="", placeholder="напр. Ивентим БГ ООД")
    reporting_year = st.number_input("Отчетна година", min_value=2000, max_value=2099,
                                     value=2024, step=1)
    amounts_in = st.selectbox("Суми в отчетите", ["хиляди лева", "лева"])
    divisor = 1000.0 if amounts_in == "хиляди лева" else 1.0

    st.divider()
    st.header("💾 Запазен мапинг")
    mapping_files = list_mapping_files()
    if mapping_files:
        chosen_file = st.selectbox("Зареди съществуващ",
                                   ["(нов)"] + [f.stem for f in mapping_files])
    else:
        chosen_file = "(нов)"
        st.caption("Нямa запазени мапинги.")

    st.divider()
    st.caption("v1.0  •  БГ Финансови Отчети")

# ── Step 1: Upload ──────────────────────────────────────────────────────────
st.markdown('<p class="step-header">Стъпка 1 — Качи оборотна ведомост</p>', unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Excel (.xlsx/.xls) или CSV файл",
    type=["xlsx", "xls", "csv"],
    help="Файлът трябва да съдържа колони за номер на сметка, наименование и крайни салда.",
)

tb_raw = None
if uploaded:
    try:
        if uploaded.name.endswith(".csv"):
            tb_raw = pd.read_csv(uploaded, encoding="utf-8-sig")
        else:
            # Try to find the right sheet
            xl = pd.ExcelFile(uploaded)
            sheet_names = xl.sheet_names
            # Heuristic: pick sheet whose name contains "об" or "TB" or first sheet
            chosen_sheet = sheet_names[0]
            for sn in sheet_names:
                low = sn.lower()
                if any(kw in low for kw in ["об", "tb", "trial", "обор", "ведом"]):
                    chosen_sheet = sn
                    break
            if len(sheet_names) > 1:
                chosen_sheet = st.selectbox("Избери лист", sheet_names,
                                             index=sheet_names.index(chosen_sheet))
            # Skip rows until we find a proper header (search first 10 rows)
            preview = pd.read_excel(uploaded, sheet_name=chosen_sheet, header=None, nrows=15)
            header_row = 0
            for i, row in preview.iterrows():
                non_empty = row.dropna()
                if len(non_empty) >= 4:
                    header_row = i
                    break
            tb_raw = pd.read_excel(uploaded, sheet_name=chosen_sheet,
                                   header=header_row)
        # Drop completely empty rows
        tb_raw = tb_raw.dropna(how="all").reset_index(drop=True)
        st.success(f"✅ Заредени {len(tb_raw)} реда, {len(tb_raw.columns)} колони.")
        with st.expander("Преглед на данните (първи 10 реда)"):
            st.dataframe(tb_raw.head(10), use_container_width=True)
    except Exception as e:
        st.error(f"Грешка при четене на файла: {e}")
        tb_raw = None

# ── Step 2: Column mapping ──────────────────────────────────────────────────
col_map = {}
cols_ready = False   # always defined; set to True below only when conditions met
if tb_raw is not None:
    st.markdown('<p class="step-header">Стъпка 2 — Свържи колоните</p>', unsafe_allow_html=True)

    cols = list(tb_raw.columns)
    auto = auto_detect_columns(cols)

    COL_LABELS = {
        "acct_num":  "Номер на сметка *",
        "acct_name": "Наименование на сметка",
        "ns_dt":     "НС Дт (начално салдо дт)",
        "ns_kt":     "НС Кт (начално салдо кт)",
        "ob_dt":     "Об Дт (обороти дт)",
        "ob_kt":     "Об Кт (обороти кт)",
        "ks_dt":     "КС Дт (крайно салдо дт) *",
        "ks_kt":     "КС Кт (крайно салдо кт) *",
    }

    none_option = ["— не е наличен —"]
    col_options = none_option + cols

    c1, c2 = st.columns(2)
    cols_left  = ["acct_num", "acct_name", "ns_dt", "ns_kt"]
    cols_right = ["ob_dt", "ob_kt", "ks_dt", "ks_kt"]

    def _select(key, container):
        detected = auto.get(key)
        idx = (col_options.index(detected)
               if detected and detected in col_options else 0)
        return container.selectbox(COL_LABELS[key], col_options,
                                   index=idx, key=f"col_{key}")

    with c1:
        for k in cols_left:
            col_map[k] = _select(k, c1)
    with c2:
        for k in cols_right:
            col_map[k] = _select(k, c2)

    required = [col_map.get("acct_num"), col_map.get("ks_dt"), col_map.get("ks_kt")]
    cols_ready = all(v and v != "— не е наличен —" for v in required)
    if not cols_ready:
        st.markdown('<div class="warn-box">⚠️ Моля, избери поне: <b>Номер на сметка</b>, <b>КС Дт</b> и <b>КС Кт</b>.</div>',
                    unsafe_allow_html=True)

# ── Step 3: Account mapping ──────────────────────────────────────────────────
mapping_df = None
if tb_raw is not None and cols_ready:
    st.markdown('<p class="step-header">Стъпка 3 — Преглед и корекция на мапи́нга</p>',
                unsafe_allow_html=True)

    # Build account list from TB
    acct_num_col  = col_map["acct_num"]
    acct_name_col = col_map.get("acct_name")

    accts = tb_raw[[acct_num_col]].copy()
    accts.columns = ["account"]
    accts["account"] = accts["account"].astype(str).str.strip()
    if acct_name_col and acct_name_col != "— не е наличен —":
        accts["name"] = tb_raw[acct_name_col].astype(str).str.strip()
    else:
        accts["name"] = ""

    # Deduplicate
    accts = accts.drop_duplicates(subset=["account"]).reset_index(drop=True)
    # Remove totals / group rows (non-numeric or very short)
    accts = accts[accts["account"].str.match(r"^\d{3,6}$")].reset_index(drop=True)

    # Try loading saved mapping
    pre_loaded = {}
    if chosen_file != "(нов)":
        try:
            saved = load_mapping(MAPPINGS_DIR / f"{chosen_file}.json")
            for _, r in saved.iterrows():
                pre_loaded[str(r["account"])] = r.to_dict()
        except Exception:
            pass

    # Build suggestion table (ASCII column names to avoid Streamlit encoding issues)
    rows = []
    for _, r in accts.iterrows():
        acct = r["account"]
        name = r["name"]
        if acct in pre_loaded:
            rows.append(pre_loaded[acct])
        else:
            b_code, o_code, atype, label = suggest_nsi(acct, name)
            rows.append({
                "account":  acct,
                "name":     name,
                "type":     atype,
                "bal_code": b_code or "",
                "opr_code": o_code or "",
                "include":  atype != "X",
            })
    mapping_df_init = pd.DataFrame(rows)

    # Persist the FULL init (with all columns, incl. non-editable account/name)
    # so we can recover them after st.data_editor drops disabled columns.
    st.session_state["_mapping_df_init"] = mapping_df_init.copy()

    st.info("Провери и коригирай при необходимост. "
            "**Тип**: A=актив, L=пасив/капитал, E=разход, R=приход, X=изключи. "
            "**Баланс/ОПР код**: кодът от шаблона.")

    BALANCE_CODES = sorted({r[0] for r in BALANCE_TEMPLATE
                             if not r[3] and r[0] not in ("AKTIV","PASIV","SUMA_A","SUMA_P")})
    OPR_CODES     = sorted({r[0] for r in OPR_TEMPLATE
                             if not r[3] and r[0] not in ("EBIT","EBT","ФР")})

    edited_df = st.data_editor(
        mapping_df_init,
        use_container_width=True,
        num_rows="fixed",
        column_config={
            "account":  st.column_config.TextColumn("Сметка", disabled=True, width=80),
            "name":     st.column_config.TextColumn("Наименование", disabled=True, width=240),
            "type":     st.column_config.SelectboxColumn("Тип", width=70,
                         options=["A","L","E","R","X"]),
            "bal_code": st.column_config.SelectboxColumn("Баланс код", width=100,
                         options=[""]+BALANCE_CODES),
            "opr_code": st.column_config.SelectboxColumn("ОПР код", width=100,
                         options=[""]+OPR_CODES),
            "include":  st.column_config.CheckboxColumn("Включи", width=70),
        },
        hide_index=True,
        key="mapping_editor_v2",
    )

    # Rebuild FULL mapping by merging the editable columns from edited_df
    # back onto mapping_df_init. This protects against Streamlit Cloud dropping
    # disabled columns (account, name) from the returned DataFrame.
    mapping_df = st.session_state["_mapping_df_init"].copy()
    if edited_df is not None:
        for col in ["type", "bal_code", "opr_code", "include"]:
            if col in edited_df.columns and len(edited_df) == len(mapping_df):
                try:
                    mapping_df[col] = edited_df[col].values
                except Exception:
                    pass  # keep init values on error
    st.session_state["_mapping_df"] = mapping_df.copy()

    # Save mapping button
    save_name = st.text_input("Запази маппинга под име (за бъдещо ползване)",
                              value=company_name or "mapping",
                              key="save_name_input")
    if st.button("💾 Запази мапинга"):
        to_save = st.session_state.get("_mapping_df", mapping_df)
        if to_save is not None:
            save_mapping(save_name, to_save)
            st.success(f"Запазено в mappings/{save_name}.json")

# ── Step 4: Generate reports ─────────────────────────────────────────────────
if mapping_df is not None:
    st.markdown('<p class="step-header">Стъпка 4 — Генерирай отчети</p>',
                unsafe_allow_html=True)

    if st.button("🔄 Генерирай Баланс и ОПР", type="primary"):
        with st.spinner("Изчисляване..."):
            try:
                # Use session-state cached mapping — always has the full structure
                # because we built it by merging edits onto mapping_df_init.
                safe_mdf = st.session_state.get("_mapping_df")
                if safe_mdf is None:
                    # Fallback: use mapping_df_init from session state
                    safe_mdf = st.session_state.get("_mapping_df_init", mapping_df)
                if safe_mdf is None:
                    st.error("Маппингът не е готов. Моля, качи файла отново.")
                    st.stop()

                # Defensive: ensure all required columns exist
                for col, default in [("account",""), ("type","X"),
                                     ("bal_code",""), ("opr_code",""), ("include", True)]:
                    if col not in safe_mdf.columns:
                        safe_mdf = safe_mdf.copy()
                        safe_mdf[col] = default

                # Attach internal column names to TB
                tb_work = tb_raw.copy()
                tb_work["_acct_num"] = tb_work[col_map["acct_num"]].astype(str).str.strip()

                ks_dt_col = col_map["ks_dt"]
                ks_kt_col = col_map["ks_kt"]
                ob_dt_col = col_map.get("ob_dt", "")
                ob_kt_col = col_map.get("ob_kt", "")

                # Filter to included accounts
                included = safe_mdf[safe_mdf["include"] == True].copy()

                # Compute
                bal_amounts = compute_balance(tb_work, included, ks_dt_col, ks_kt_col)
                opr_amounts = compute_opr(
                    tb_work, included,
                    ks_dt_col, ks_kt_col,
                    ob_dt_col if ob_dt_col and ob_dt_col != "— не е наличен —" else ks_dt_col,
                    ob_kt_col if ob_kt_col and ob_kt_col != "— не е наличен —" else ks_kt_col,
                )

                st.session_state["bal_amounts"] = bal_amounts
                st.session_state["opr_amounts"] = opr_amounts
                st.success("✅ Отчетите са генерирани успешно!")

            except Exception as e:
                st.error(f"Грешка: {e}")
                import traceback
                st.code(traceback.format_exc())

# ── Step 5: Display & Download ───────────────────────────────────────────────
if "bal_amounts" in st.session_state and "opr_amounts" in st.session_state:
    bal_amounts = st.session_state["bal_amounts"]
    opr_amounts = st.session_state["opr_amounts"]

    st.markdown('<p class="step-header">Стъпка 5 — Преглед и изтегляне</p>',
                unsafe_allow_html=True)

    tab_bal, tab_opr = st.tabs(["📋 Баланс", "📋 ОПР"])

    def fmt(v, div=1.0):
        if v is None or v == 0:
            return ""
        return f"{v / div:,.3f}" if abs(v) >= 0.001 else ""

    with tab_bal:
        bal_rows = []
        for code, label, indent, is_total, _sign in BALANCE_TEMPLATE:
            amount = bal_amounts.get(code, 0.0)
            bal_rows.append({
                "Показател": "  " * indent + label,
                "Код":       code if code not in ("AKTIV","PASIV","SUMA_A","SUMA_P") else "",
                f"Сума ({amounts_in})": fmt(amount, divisor),
                "_bold": is_total,
            })
        bal_display = pd.DataFrame(bal_rows)
        st.dataframe(
            bal_display[["Показател","Код",f"Сума ({amounts_in})"]],
            use_container_width=True,
            hide_index=True,
            height=800,
        )

    with tab_opr:
        opr_rows = []
        for code, label, indent, is_total, sign in OPR_TEMPLATE:
            amount = opr_amounts.get(code, 0.0)
            opr_rows.append({
                "Показател": "  " * indent + label,
                "Код":       code if code not in ("EBIT","EBT","ФР") else "",
                f"Сума ({amounts_in})": fmt(amount, divisor),
            })
        opr_display = pd.DataFrame(opr_rows)
        st.dataframe(
            opr_display[["Показател","Код",f"Сума ({amounts_in})"]],
            use_container_width=True,
            hide_index=True,
            height=600,
        )

    # Download Excel
    excel_bytes = build_excel(bal_amounts, opr_amounts, company_name)
    fname = f"Отчети_{company_name}_{reporting_year}.xlsx".replace(" ", "_")
    st.download_button(
        label="⬇️ Изтегли Excel (Баланс + ОПР)",
        data=excel_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    # Check balance
    suma_a = bal_amounts.get("SUMA_A", 0)
    suma_p = bal_amounts.get("SUMA_P", 0)
    diff = abs(suma_a - suma_p)
    if diff < 1:
        st.markdown(
            f'<div class="success-box">✅ Балансът се балансира: '
            f'Актив = {suma_a/1000:,.3f} хил. лв. = Пасив = {suma_p/1000:,.3f} хил. лв.</div>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div class="warn-box">⚠️ Дисбаланс: Актив = {suma_a/1000:,.3f}, '
            f'Пасив = {suma_p/1000:,.3f}, разлика = {diff/1000:,.3f} хил. лв. '
            f'Провери маппинга.</div>',
            unsafe_allow_html=True)
